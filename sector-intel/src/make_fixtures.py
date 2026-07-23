"""Synthesise messy DoE/QILT-shaped workbooks so the pipeline runs end to end.

These are NOT real figures. They stand in for the real government workbooks so
acquisition, the hard parsing, name resolution, benchmarking and validation can
be proven today. Real refresh = drop the real files into data/raw and rerun.
The messiness (title rows, a two-row header band, merged cells, footnotes,
'np' suppressed values, a couple of name variants) mirrors the real shapes.
"""
from __future__ import annotations
import random, zlib
from openpyxl import Workbook
from openpyxl.styles import Font
from . import config

GROUP_SCALE = {"Go8": 1.0, "ATN": 0.72, "IRU": 0.55, "RUN": 0.34, "Unaligned": 0.5}
# A few source-printed name variants to exercise resolve.py (alias + fuzzy).
NAME_VARIANTS = {
    "unsw": "University of NSW",
    "sydney": "University of Sydney ",          # trailing space
    "rmit": "RMIT",
    "uts": "UTS",
    "qut": "Queensland Uni of Technology",       # fuzzy only
}

def _rng(iid, salt):
    # stable seed (str.__hash__ is per-process randomised) so fixtures are deterministic
    return random.Random(zlib.crc32(f"{iid}:{salt}".encode()))

def _profile(inst):
    g = inst["mission_group"]
    r = _rng(inst["institution_id"], "base")
    scale = GROUP_SCALE.get(g, 0.5) * r.uniform(0.8, 1.2)
    return dict(
        head=int(max(3500, 62000 * scale)),
        intl={"Go8":34,"ATN":26,"IRU":20,"RUN":11,"Unaligned":22}[g] + r.uniform(-6,6),
        online={"RUN":46,"Unaligned":24,"IRU":22,"ATN":20,"Go8":9}[g] + r.uniform(-7,10),
        attr={"Go8":11,"ATN":15,"IRU":18,"RUN":21,"Unaligned":16}[g] + r.uniform(-3,4),
        exp={"Go8":78,"ATN":76,"IRU":75,"RUN":77,"Unaligned":76}[g] + r.uniform(-5,5),
        grad={"Go8":80,"ATN":76,"IRU":73,"RUN":72,"Unaligned":74}[g] + r.uniform(-6,6),
        revenue=round(max(120, 2600 * scale)),
        intl_rev=min(55, max(5, ({"Go8":34,"ATN":26,"IRU":20,"RUN":11,"Unaligned":22}[g]) * 0.9)),
        staff=None,
    )

def _yfac(i): return 0.94 + 0.03 * i

def _printed_name(inst):
    return NAME_VARIANTS.get(inst["institution_id"], inst["institution_name"])

def _bold(ws, cell):
    ws[cell].font = Font(bold=True)

def build():
    ref = config.load_reference()
    insts = list(ref["institutions"].values())
    import yaml
    years = yaml.safe_load(config.MANIFEST.read_text())["fixture_years"]
    (config.RAW / "doe").mkdir(parents=True, exist_ok=True)
    (config.RAW / "qilt").mkdir(parents=True, exist_ok=True)
    log = config.load_logger()

    for yi, year in enumerate(years):
        yf = _yfac(yi)
        # ---- DoE all students ----
        wb = Workbook(); ws = wb.active; ws.title = "Table 2.1 All students"
        ws["A1"] = f"Table 2.1: All students by provider, {year}"; _bold(ws, "A1")
        ws["A2"] = "Selected Higher Education Statistics — Student data"
        ws["A4"] = "Provider"; ws["B4"] = "All students (no.)"; ws["C4"] = "EFTSL"
        ws["D4"] = "Commencing (no.)"; ws["E4"] = "International (%)"
        ws["F4"] = "External/online load (%)"; ws["G4"] = "First-year attrition (%)"
        for c in "ABCDEFG": _bold(ws, f"{c}4")
        row = 5
        for inst in insts:
            p = _profile(inst); r = _rng(inst["institution_id"], f"doe{year}")
            head = int(p["head"] * yf)
            ws.cell(row, 1, _printed_name(inst))
            ws.cell(row, 2, head)
            ws.cell(row, 3, round(head * 0.62))
            ws.cell(row, 4, round(head * 0.36))
            ws.cell(row, 5, round(max(4, min(48, p["intl"])), 1))
            ws.cell(row, 6, round(max(3, min(62, p["online"] * (0.96 + 0.06 * yf))), 1))
            # sprinkle a few suppressed attrition values
            ws.cell(row, 7, "np" if r.random() < 0.05 else round(max(6, min(26, p["attr"] * (1.03 - 0.03 * yf))), 1))
            row += 1
        ws.cell(row + 1, 1, "np = not published (below reporting threshold)")
        wb.save(config.RAW / "doe" / f"doe_student_allstudents_{year}.xlsx")

        # ---- DoE staff + finance ----
        wb = Workbook(); ws = wb.active; ws.title = "Table 4 Staff and finance"
        ws["A1"] = f"Table 4: Staff (FTE) and finance by provider, {year}"; _bold(ws, "A1")
        ws["A3"] = "Provider"; ws["B3"] = "Staff FTE"; ws["C3"] = "Total revenue ($m)"; ws["D3"] = "International fee share (%)"
        for c in "ABCD": _bold(ws, f"{c}3")
        row = 4
        for inst in insts:
            p = _profile(inst); head = int(p["head"] * yf)
            ws.cell(row, 1, _printed_name(inst))
            ws.cell(row, 2, round(head * 0.09))
            ws.cell(row, 3, round(p["revenue"] * yf))
            ws.cell(row, 4, round(max(4, min(56, p["intl_rev"] * (0.97 + 0.03 * yf))), 1))
            row += 1
        wb.save(config.RAW / "doe" / f"doe_staff_finance_{year}.xlsx")

        # ---- QILT SES (coded sheet) ----
        wb = Workbook(); ws = wb.active; ws.title = "STMT_SES_ALL_1Y"
        ws["A1"] = f"Student Experience Survey — national tables, {year}"; _bold(ws, "A1")
        ws["A2"] = "Undergraduate, percent positive"
        ws["A4"] = "Institution"; ws["B4"] = "Overall experience (%)"; ws["C4"] = "Teaching quality (%)"
        ws["D4"] = "Student support (%)"; ws["E4"] = "Graduate employment (%)"
        for c in "ABCDE": _bold(ws, f"{c}4")
        row = 5
        for inst in insts:
            p = _profile(inst); r = _rng(inst["institution_id"], f"ses{year}")
            exp = max(60, min(88, p["exp"] + (yf - 1) * 4))
            ws.cell(row, 1, _printed_name(inst))
            ws.cell(row, 2, round(exp, 1))
            ws.cell(row, 3, round(min(88, exp + r.uniform(-2, 3)), 1))
            ws.cell(row, 4, round(max(58, exp - r.uniform(2, 8)), 1))
            ws.cell(row, 5, round(max(55, min(88, p["grad"] + (yf - 1) * 6)), 1))
            row += 1
        wb.save(config.RAW / "qilt" / f"qilt_ses_{year}.xlsx")
        log.info(f"fixtures written for {year}")
    return years

if __name__ == "__main__":
    build()
