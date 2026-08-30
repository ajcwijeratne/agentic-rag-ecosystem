"""Turn real DoE/QILT workbooks into tidy per-source CSVs.

Government Excel is built for human reading, not machines: title rows, header
bands spanning two rows, merged cells, footnotes below the data, suppressed
'np' cells, institutions as columns instead of rows, and layouts that drift
year to year (sometimes even the sheet a table lives on renumbers). Each real
source gets its own small parser function below rather than one generic
config-driven reader -- these workbooks are too irregular for that to hold up,
and a layout change next year is now a targeted edit to one function, guided
by the assertions/lookups failing loudly, not a silent mismatch.

Output is unchanged: one tidy CSV per parsed source, columns
source_name, metric_id, year, value, suppressed.
"""
from __future__ import annotations
import csv, re, zipfile
from xml.etree import ElementTree as ET
from openpyxl import load_workbook
from . import config

_NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _clean(v):
    """Plain numeric cell -> (value, suppressed)."""
    if v is None:
        return None, False
    s = str(v).strip()
    if s.lower() in config.SUPPRESSED:
        return None, True
    try:
        return float(s.replace(",", "").replace("$", "").replace("%", "")), False
    except ValueError:
        return None, False


def _clean_ci(v):
    """QILT '84.1 (83.6, 84.6)' cell -> (point estimate, suppressed)."""
    if v is None:
        return None, False
    s = str(v).strip()
    if not s or s.lower() in config.SUPPRESSED:
        return None, True
    m = re.match(r"^(-?[\d.]+)", s)
    if not m:
        return None, False
    try:
        return float(m.group(1)), False
    except ValueError:
        return None, False


def _norm_header(h):
    if h is None:
        return ""
    h = str(h).replace("\xa0", " ")
    return re.sub(r"\s+", " ", h).strip()


def _strip_footnote(h):
    """'Permanent resident(4.17)' / 'Non-award Courses /Microcredentials(a)' -> 'Permanent resident'."""
    h = _norm_header(h)
    while True:
        new = re.sub(r"\s*\([^)]*\)\s*$", "", h).strip()
        new = re.sub(r"\s*/[^/()]*$", "", new).strip() if new == h else new
        if new == h:
            break
        h = new
    return h


# ---------------------------------------------------------------------------
# QILT raw-XML reader. openpyxl's read_only mode mis-reads these workbooks:
# the exported <dimension> tag is stuck at 'A1' even though real data runs for
# dozens of rows, so row iteration silently stops after row 1 (no error). The
# workbook has no pivot cache either -- the fix is to read the sheet XML
# directly and ignore the declared dimension.
# ---------------------------------------------------------------------------
def _qilt_shared_strings(z):
    xml = z.read("xl/sharedStrings.xml").decode("utf-8", errors="replace")
    root = ET.fromstring(xml)
    return ["".join((t.text or "") for t in si.findall(".//a:t", _NS))
            for si in root.findall("a:si", _NS)]


def _qilt_sheet_path(z, sheet_name):
    wbxml = z.read("xl/workbook.xml").decode("utf-8", errors="replace")
    rid = dict(re.findall(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', wbxml)).get(sheet_name)
    if rid is None:
        return None
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="replace")
    return dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels)).get(rid)


def _qilt_read_sheet(path, sheet_name) -> list[dict]:
    """One dict per row: {column_letter: value}, in row order, 0-indexed list position."""
    with zipfile.ZipFile(path) as z:
        shared = _qilt_shared_strings(z)
        sp = _qilt_sheet_path(z, sheet_name)
        if not sp:
            return []
        xml = z.read(f"xl/{sp}").decode("utf-8", errors="replace")
        root = ET.fromstring(xml)
        rows = []
        for row_el in root.find("a:sheetData", _NS).findall("a:row", _NS):
            rv = {}
            for c in row_el.findall("a:c", _NS):
                col = re.match(r"([A-Z]+)", c.attrib["r"]).group(1)
                t = c.attrib.get("t", "")
                v = c.find("a:v", _NS)
                rv[col] = None if v is None else (shared[int(v.text)] if t == "s" else v.text)
            rows.append(rv)
        return rows


# ---------------------------------------------------------------------------
# doe_students: Section 2 "All Students" workbook (A01, A04, A05).
# Table 2.1 (the fixture-era 'sheet') has NO institution rows -- it's an age
# group breakdown -- so real figures come from three different tables. The
# citizenship table's sheet name shifts in 2024.
# ---------------------------------------------------------------------------
_DOE_STUDENTS_CITIZENSHIP_SHEET = {2022: "2.10", 2023: "2.10", 2024: "2.12"}
# DoE tables include non-institution subtotal rows mixed in with real
# institutions -- these are the two seen in the real files, excluded rather
# than left for resolve.py to bin as unmatched noise.
_DOE_STUDENTS_AGGREGATE_ROWS = {
    "non-university higher education institutions",
    "private universities (table c) and non-university higher education institutions",
}


def _parse_doe_students(src, raw, log):
    year = src["release_year"]
    wb = load_workbook(raw, data_only=True, read_only=True)
    out = []

    # A01 total enrolments: Table 2.5, TOTAL column, by institution.
    ws = wb["2.5"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [_norm_header(c) for c in rows[2]]
    name_i = hdr.index("Institution")
    total_i = next(i for i, h in enumerate(hdr) if h.lower() == "total")
    for r in rows[3:]:
        if not r or r[name_i] is None:
            continue
        name = str(r[name_i]).strip()
        if not name or name.lower() in _DOE_STUDENTS_AGGREGATE_ROWS:
            continue
        val, supp = _clean(r[total_i] if total_i < len(r) else None)
        if val is not None or supp:
            out.append({"source_name": name, "metric_id": "A01", "year": year,
                        "value": val, "suppressed": int(supp)})

    # A05 online/external share: Table 2.7, External / (Internal+External+Multi-modal).
    # 'Not provided' and the mode subtotal are deliberately left out of the
    # denominator -- three different columns are literally named 'TOTAL' in the
    # 2022/2023 layout (mode/type/gender subtotals) and Internal+External+
    # Multi-modal already accounts for everything but rounding.
    ws = wb["2.7"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = 3 if year == 2024 else 4   # 2024 dropped the Type/Gender columns and the group row
    hdr = [_norm_header(c) for c in rows[header_row - 1]]
    name_i = hdr.index("Institution")
    int_i, ext_i, multi_i = hdr.index("Internal"), hdr.index("External"), hdr.index("Multi-modal")
    for r in rows[header_row:]:
        if not r or r[name_i] is None:
            continue
        name = str(r[name_i]).strip()
        if not name or name.lower() in _DOE_STUDENTS_AGGREGATE_ROWS:
            continue
        vi, _ = _clean(r[int_i] if int_i < len(r) else None)
        ve, se = _clean(r[ext_i] if ext_i < len(r) else None)
        vm, _ = _clean(r[multi_i] if multi_i < len(r) else None)
        denom = (vi or 0) + (ve or 0) + (vm or 0)
        if denom:
            out.append({"source_name": name, "metric_id": "A05", "year": year,
                        "value": round(100 * (ve or 0) / denom, 2), "suppressed": 0})
        elif se:
            out.append({"source_name": name, "metric_id": "A05", "year": year,
                        "value": None, "suppressed": 1})

    # A04 international enrolment share: citizenship table, computed as
    # Total minus the (stable-named) domestic categories minus Not provided --
    # the overseas column wording itself changes every year, so naming it
    # directly would break on the next DoE release.
    sheet_name = _DOE_STUDENTS_CITIZENSHIP_SHEET[year]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    group_row = [_strip_footnote(c) for c in rows[2]]
    sub_row = [_strip_footnote(c) for c in rows[3]]
    width = max(len(group_row), len(sub_row))
    group_row += [""] * (width - len(group_row))
    sub_row += [""] * (width - len(sub_row))
    hdr = [sub_row[i] or group_row[i] for i in range(width)]
    name_i = hdr.index("Institution")
    total_i = next(i for i, h in enumerate(hdr) if h.lower() == "total")
    domestic_labels = {"australian citizen", "new zealand citizen", "permanent resident",
                        "permanent humanitarian visa", "not provided"}
    dom_idx = [i for i, h in enumerate(hdr) if h.lower() in domestic_labels]
    for r in rows[4:]:
        if not r or name_i >= len(r) or r[name_i] is None:
            continue
        name = str(r[name_i]).strip()
        if not name or name.lower() in _DOE_STUDENTS_AGGREGATE_ROWS:
            continue
        total, tsupp = _clean(r[total_i] if total_i < len(r) else None)
        if total is None:
            if tsupp:
                out.append({"source_name": name, "metric_id": "A04", "year": year,
                            "value": None, "suppressed": 1})
            continue
        dom_sum = sum((_clean(r[i] if i < len(r) else None)[0] or 0) for i in dom_idx)
        out.append({"source_name": name, "metric_id": "A04", "year": year,
                    "value": round(100 * max(0.0, total - dom_sum) / total, 2), "suppressed": 0})
    return out


# ---------------------------------------------------------------------------
# doe_finance: Finance Tables workbook, sheet "Financial Performance" (D01, D03).
# Institutions are COLUMNS here, not rows -- the header row holds institution
# names and each subsequent row is one revenue/expense line item. Values are
# in $'000. Row labels and even whether there's a second alt-label column
# differ by year, so both the header row and the target rows are located by
# searching, not by fixed row numbers.
# ---------------------------------------------------------------------------
def _parse_doe_finance(src, raw, log):
    year = src["release_year"]
    wb = load_workbook(raw, data_only=True, read_only=True)
    ws = wb["Financial Performance"]
    rows = list(ws.iter_rows(values_only=True))

    hdr_i = next((i for i, r in enumerate(rows)
                  if r and "Macquarie University" in r and "Charles Sturt University" in r), None)
    if hdr_i is None:
        log.error(f"parse FAIL {src['key']}: institution header row not found in Financial Performance")
        return []
    hdr = rows[hdr_i]
    insts = {i: str(h).strip() for i, h in enumerate(hdr)
             if h and not str(h).strip().startswith("All Institutions")}

    total_i = next((i for i, r in enumerate(rows) if r and r[0]
                     and str(r[0]).strip().startswith("Total Revenues from Continuing Operations")), None)
    intl_i = next((i for i, r in enumerate(rows) if r and r[0]
                    and str(r[0]).strip() in ("International Students", "Fee Paying Overseas Students")), None)
    if total_i is None:
        log.error(f"parse FAIL {src['key']}: 'Total Revenues from Continuing Operations' row not found")
        return []
    if intl_i is None:
        log.warning(f"parse {src['key']}: international-fee row not found; D03 skipped for this year")

    total_row = rows[total_i]
    intl_row = rows[intl_i] if intl_i is not None else None
    out = []
    for i, name in insts.items():
        tv, tsupp = _clean(total_row[i] if i < len(total_row) else None)
        if tv is not None:
            out.append({"source_name": name, "metric_id": "D01", "year": year,
                        "value": round(tv / 1000, 2), "suppressed": 0})   # $'000 -> $m
        elif tsupp:
            out.append({"source_name": name, "metric_id": "D01", "year": year,
                        "value": None, "suppressed": 1})
        if intl_row is None:
            continue
        iv, isupp = _clean(intl_row[i] if i < len(intl_row) else None)
        if tv:
            out.append({"source_name": name, "metric_id": "D03", "year": year,
                        "value": round(100 * (iv or 0) / tv, 2), "suppressed": 0})
        elif isupp or tsupp:
            out.append({"source_name": name, "metric_id": "D03", "year": year,
                        "value": None, "suppressed": 1})
    return out


# ---------------------------------------------------------------------------
# doe_staff: Staff Appendix 1 workbook (C01). DoE renumbers this appendix's
# tables most years, so the sheet holding the by-institution Published-vs-
# Actual comparison is looked up per year rather than assumed fixed.
# ---------------------------------------------------------------------------
_DOE_STAFF_SHEET = {2022: "2", 2023: "A1.2", 2024: "A1.6"}
_DOE_STAFF_NAME_COL = {2022: 1, 2023: 2, 2024: 2}   # 2023/2024 insert an institution-code column


def _parse_doe_staff(src, raw, log):
    year = src["release_year"]
    wb = load_workbook(raw, data_only=True, read_only=True)
    sheet = _DOE_STAFF_SHEET[year]
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [_norm_header(c) for c in rows[3]]
    name_i = _DOE_STAFF_NAME_COL[year]
    try:
        fte_i = hdr.index("Total FTE including Actual Casual FTE")
    except ValueError:
        log.error(f"parse FAIL {src['key']}: 'Total FTE including Actual Casual FTE' column "
                  f"not found on sheet {sheet!r} — header may have shifted, header was {hdr}")
        return []
    out = []
    for r in rows[4:]:
        if not r or name_i >= len(r) or r[name_i] is None:
            continue
        name = str(r[name_i]).strip()
        if not name:
            continue
        val, supp = _clean(r[fte_i] if fte_i < len(r) else None)
        if val is not None or supp:
            out.append({"source_name": name, "metric_id": "C01", "year": year,
                        "value": round(val, 1) if val is not None else None, "suppressed": int(supp)})
    return out


# ---------------------------------------------------------------------------
# qilt_ses: SES national tables, sheet FOCUS_UG_UNI_1Y_INST_CI (B01, B02, B04).
# QILT redesigned the SES scales for 2023 (Teaching Quality + Engagement
# merged into one scale; Student Support + Services likewise), so the column
# headers genuinely differ 2022 vs 2023+. Matched by header substring, not
# position, so that's absorbed automatically. B06 (graduate employment) is not
# an SES metric at all -- it's Graduate Outcomes Survey -- and is not parsed
# here; no GOS source is configured yet.
# ---------------------------------------------------------------------------
_QILT_SHEET = "FOCUS_UG_UNI_1Y_INST_CI"
_QILT_COL_MATCH = {
    "B01": "quality of entire educational",
    "B02": "teaching quality",
    "B04": "student support",
}


def _parse_qilt_ses(src, raw, log):
    year = src["release_year"]
    rows = _qilt_read_sheet(raw, _QILT_SHEET)
    if not rows or len(rows) < 5:
        log.error(f"parse FAIL {src['key']}: sheet {_QILT_SHEET!r} not found or empty")
        return []
    hdr_row = rows[3]

    def find_col(substr):
        for col, val in hdr_row.items():
            if val and substr in _norm_header(val).lower():
                return col
        return None

    col_map = {mid: find_col(substr) for mid, substr in _QILT_COL_MATCH.items()}
    missing = [mid for mid, col in col_map.items() if col is None]
    if missing:
        log.warning(f"parse {src['key']}: column(s) not found for {missing} — "
                    f"header row was {hdr_row}")

    name_col = "B"
    out = []
    for r in rows[4:]:
        name = (r.get(name_col) or "").strip()
        if not name or name.lower().startswith("all universit"):
            break
        for mid, col in col_map.items():
            if col is None:
                continue
            val, supp = _clean_ci(r.get(col))
            if val is not None or supp:
                out.append({"source_name": name, "metric_id": mid, "year": year,
                            "value": val, "suppressed": int(supp)})
    return out


PARSER_FUNCS = {
    "doe_students": _parse_doe_students,
    "doe_finance": _parse_doe_finance,
    "doe_staff": _parse_doe_staff,
    "qilt_ses": _parse_qilt_ses,
}


def parse_source(src: dict) -> list[dict]:
    log = config.load_logger()
    raw = config.ROOT / src["raw_path"]
    if not raw.exists():
        log.warning(f"parse skip (missing raw): {src['key']}")
        return []
    fn = PARSER_FUNCS.get(src["parser"])
    if not fn:
        log.error(f"parse FAIL {src['key']}: no parser registered for {src['parser']!r}")
        return []
    try:
        out = fn(src, raw, log)
    except Exception as e:  # noqa: BLE001
        log.error(f"parse FAIL {src['key']}: {type(e).__name__}: {e}")
        return []
    dest = config.PARSED / f"{src['key']}.csv"
    with open(dest, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["source_name", "metric_id", "year", "value", "suppressed"])
        w.writeheader(); w.writerows(out)
    log.info(f"parsed {src['key']}: {len(out)} observations -> {dest.name}")
    return out


def run(sources: list[dict]) -> list[dict]:
    allrows = []
    for src in sources:
        allrows += parse_source(src)
    return allrows


if __name__ == "__main__":
    import yaml
    man = yaml.safe_load(config.MANIFEST.read_text())
    run(man["sources"])
